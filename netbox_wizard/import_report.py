"""
Import Report Generation for NetBox Import Wizard
Generates CSV and Excel reports for administrative follow-up
"""
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from PyQt6.QtCore import QStandardPaths

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ImportReportGenerator:
    """Generate detailed reports for imported devices"""

    def __init__(self):
        self.import_results = []
        self.netbox_data = {}
        self.topology_file = ""
        self.import_timestamp = None

    def set_import_data(self, results: List[Dict], netbox_data: Dict, topology_file: str = ""):
        """Set the import results and supporting data"""
        self.import_results = results
        self.netbox_data = netbox_data
        self.topology_file = topology_file
        self.import_timestamp = datetime.now()

    def generate_csv_report(self, parent_widget=None) -> bool:
        """Generate CSV report of import results"""
        if not self.import_results:
            QMessageBox.information(parent_widget, "No Data", "No import results to report")
            return False

        # Get save location
        documents_dir = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DocumentsLocation)
        timestamp = self.import_timestamp.strftime("%Y%m%d_%H%M%S")
        default_filename = f"netbox_import_report_{timestamp}.csv"
        default_path = str(Path(documents_dir) / default_filename)

        file_path, _ = QFileDialog.getSaveFileName(
            parent_widget,
            "Save Import Report",
            default_path,
            "CSV files (*.csv)"
        )

        if not file_path:
            return False

        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)

                # Write header information
                writer.writerow(['NetBox Import Report'])
                writer.writerow(['Generated:', self.import_timestamp.strftime("%Y-%m-%d %H:%M:%S")])
                writer.writerow(['Topology File:', self.topology_file or 'Unknown'])
                writer.writerow([])  # Empty row

                # Write summary
                successful = len([r for r in self.import_results if r.get('success', False)])
                failed = len(self.import_results) - successful

                writer.writerow(['Import Summary'])
                writer.writerow(['Total Devices Processed:', len(self.import_results)])
                writer.writerow(['Successfully Created:', successful])
                writer.writerow(['Failed to Create:', failed])
                writer.writerow(['Success Rate:',
                                 f"{(successful / len(self.import_results) * 100):.1f}%" if self.import_results else "0%"])
                writer.writerow([])  # Empty row

                # Write detailed results
                writer.writerow([
                    'Device Name', 'Status', 'NetBox ID', 'IP Address', 'Platform',
                    'Site', 'Role', 'Device Type', 'Message', 'Admin Notes'
                ])

                for result in self.import_results:
                    writer.writerow([
                        result.get('name', ''),
                        'SUCCESS' if result.get('success', False) else 'FAILED',
                        result.get('netbox_id', ''),
                        result.get('ip_address', ''),
                        result.get('platform_name', ''),
                        result.get('site_name', ''),
                        result.get('role_name', ''),
                        result.get('device_type_name', ''),
                        result.get('message', ''),
                        ''  # Empty admin notes column for manual entry
                    ])

            QMessageBox.information(
                parent_widget,
                "Report Generated",
                f"Import report saved to:\n{file_path}\n\n{len(self.import_results)} devices reported"
            )
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "Report Error", f"Failed to generate report:\n{str(e)}")
            return False

    def generate_excel_report(self, parent_widget=None) -> bool:
        """Generate comprehensive Excel report with admin follow-up template"""
        if not EXCEL_AVAILABLE:
            QMessageBox.warning(
                parent_widget,
                "Excel Not Available",
                "openpyxl library not installed. Please install it for Excel reports:\npip install openpyxl"
            )
            return False

        if not self.import_results:
            QMessageBox.information(parent_widget, "No Data", "No import results to report")
            return False

        # Get save location
        documents_dir = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DocumentsLocation)
        timestamp = self.import_timestamp.strftime("%Y%m%d_%H%M%S")
        default_filename = f"netbox_import_admin_report_{timestamp}.xlsx"
        default_path = str(Path(documents_dir) / default_filename)

        file_path, _ = QFileDialog.getSaveFileName(
            parent_widget,
            "Save Import Report (Excel)",
            default_path,
            "Excel files (*.xlsx)"
        )

        if not file_path:
            return False

        try:
            workbook = openpyxl.Workbook()

            # Create summary worksheet
            self._create_summary_sheet(workbook)

            # Create detailed results worksheet
            self._create_details_sheet(workbook)

            # Create admin follow-up worksheet
            self._create_admin_template_sheet(workbook)

            # Save workbook
            workbook.save(file_path)

            QMessageBox.information(
                parent_widget,
                "Excel Report Generated",
                f"Comprehensive import report saved to:\n{file_path}\n\nIncludes admin follow-up template"
            )
            return True

        except Exception as e:
            QMessageBox.critical(parent_widget, "Report Error", f"Failed to generate Excel report:\n{str(e)}")
            return False

    def _create_summary_sheet(self, workbook):
        """Create summary worksheet"""
        ws = workbook.active
        ws.title = "Import Summary"

        # Header styling
        header_font = Font(bold=True, size=14)

        # Title
        ws['A1'] = "NetBox Import Report"
        ws['A1'].font = Font(bold=True, size=16)

        # Basic info
        ws['A3'] = "Import Date:"
        ws['B3'] = self.import_timestamp.strftime("%Y-%m-%d %H:%M:%S")
        ws['A4'] = "Topology File:"
        ws['B4'] = self.topology_file or 'Unknown'

        # Summary statistics
        successful = len([r for r in self.import_results if r.get('success', False)])
        failed = len(self.import_results) - successful

        ws['A6'] = "Summary Statistics"
        ws['A6'].font = header_font

        ws['A7'] = "Total Devices:"
        ws['B7'] = len(self.import_results)
        ws['A8'] = "Successfully Created:"
        ws['B8'] = successful
        ws['A9'] = "Failed:"
        ws['B9'] = failed
        ws['A10'] = "Success Rate:"
        ws['B10'] = f"{(successful / len(self.import_results) * 100):.1f}%" if self.import_results else "0%"

        # Platform breakdown
        platform_counts = {}
        for result in self.import_results:
            platform = result.get('platform_name', 'Unknown')
            platform_counts[platform] = platform_counts.get(platform, 0) + 1

        ws['A12'] = "Platform Breakdown"
        ws['A12'].font = header_font

        row = 13
        for platform, count in sorted(platform_counts.items()):
            ws[f'A{row}'] = platform
            ws[f'B{row}'] = count
            row += 1

    def _create_details_sheet(self, workbook):
        """Create detailed results worksheet"""
        ws = workbook.create_sheet("Import Details")

        # Headers
        headers = [
            'Device Name', 'Status', 'NetBox ID', 'IP Address', 'Platform',
            'Site', 'Role', 'Device Type', 'Import Message'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Data rows
        for row, result in enumerate(self.import_results, 2):
            ws.cell(row=row, column=1, value=result.get('name', ''))
            ws.cell(row=row, column=2, value='SUCCESS' if result.get('success', False) else 'FAILED')
            ws.cell(row=row, column=3, value=result.get('netbox_id', ''))
            ws.cell(row=row, column=4, value=result.get('ip_address', ''))
            ws.cell(row=row, column=5, value=result.get('platform_name', ''))
            ws.cell(row=row, column=6, value=result.get('site_name', ''))
            ws.cell(row=row, column=7, value=result.get('role_name', ''))
            ws.cell(row=row, column=8, value=result.get('device_type_name', ''))
            ws.cell(row=row, column=9, value=result.get('message', ''))

            # Color code status
            status_cell = ws.cell(row=row, column=2)
            if result.get('success', False):
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_admin_template_sheet(self, workbook):
        """Create admin follow-up template worksheet"""
        ws = workbook.create_sheet("Admin Follow-up")

        # Instructions
        ws['A1'] = "NetBox Administrative Follow-up Template"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = "Instructions:"
        ws['A4'] = "1. Complete the missing information for successfully imported devices"
        ws['A5'] = "2. Assign devices to racks and locations as needed"
        ws['A6'] = "3. Add asset tags, serial numbers, and custom field values"
        ws['A7'] = "4. Configure device relationships and connections"
        ws['A8'] = ""

        # Headers for follow-up table
        headers = [
            'Device Name', 'NetBox ID', 'Status', 'Rack Assignment', 'Position',
            'Asset Tag', 'Serial Number', 'Location', 'Custom Fields', 'Notes'
        ]

        start_row = 10
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # Add successful devices for follow-up
        successful_devices = [r for r in self.import_results if r.get('success', False)]

        for row, result in enumerate(successful_devices, start_row + 1):
            ws.cell(row=row, column=1, value=result.get('name', ''))
            ws.cell(row=row, column=2, value=result.get('netbox_id', ''))
            ws.cell(row=row, column=3, value='Imported - Needs Configuration')
            # Columns 4-10 left blank for admin input

            # Light yellow background for data entry rows
            for col in range(4, 11):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width